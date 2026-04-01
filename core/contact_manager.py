"""
ContactManager - 연락처 관리 모듈
엑셀 업로드, 카테고리 관리, CRUD
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None


class Contact:
    """연락처 데이터 클래스"""

    def __init__(self, name: str, category: str = "other", phone: str = "",
                 company: str = "", position: str = "", memo: str = "",
                 birthday: str = "", anniversary: str = "",
                 contact_id: str = None):
        self.id = contact_id or f"{name}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        self.name = name
        self.category = category
        self.phone = phone
        self.company = company
        self.position = position
        self.memo = memo
        self.birthday = birthday      # MM-DD 형식
        self.anniversary = anniversary  # MM-DD 형식
        self.created_at = datetime.now().isoformat()
        self.last_sent = None
        self.send_count = 0

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "name": self.name,
            "category": self.category,
            "phone": self.phone,
            "company": self.company,
            "position": self.position,
            "memo": self.memo,
            "birthday": self.birthday,
            "anniversary": self.anniversary,
            "created_at": self.created_at,
            "last_sent": self.last_sent,
            "send_count": self.send_count
        }

    @classmethod
    def from_dict(cls, data: dict) -> "Contact":
        contact = cls(
            name=data["name"],
            category=data.get("category", "other"),
            phone=data.get("phone", ""),
            company=data.get("company", ""),
            position=data.get("position", ""),
            memo=data.get("memo", ""),
            birthday=data.get("birthday", ""),
            anniversary=data.get("anniversary", ""),
            contact_id=data.get("id")
        )
        contact.created_at = data.get("created_at", contact.created_at)
        contact.last_sent = data.get("last_sent")
        contact.send_count = data.get("send_count", 0)
        return contact


class ContactManager:
    """연락처 CRUD 및 관리"""

    DEFAULT_CATEGORIES = ["friend", "family", "business", "vip", "other"]

    def __init__(self, data_path: str = None):
        self.data_path = Path(data_path) if data_path else Path("./data/contacts.json")
        self.contacts: list[Contact] = []
        self.custom_categories: list[str] = []
        self.load()

    def load(self):
        """JSON 파일에서 연락처 로드"""
        if self.data_path.exists():
            try:
                with open(self.data_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.contacts = [
                    Contact.from_dict(c) for c in data.get("contacts", [])
                ]
                self.custom_categories = data.get("custom_categories", [])
            except (json.JSONDecodeError, KeyError):
                self.contacts = []

    def save(self):
        """연락처를 JSON 파일로 저장"""
        self.data_path.parent.mkdir(parents=True, exist_ok=True)
        data = {
            "contacts": [c.to_dict() for c in self.contacts],
            "custom_categories": self.custom_categories,
            "last_updated": datetime.now().isoformat()
        }
        with open(self.data_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def add(self, contact: Contact) -> bool:
        """연락처 추가"""
        # 중복 확인 (이름 기준)
        for c in self.contacts:
            if c.name == contact.name and c.category == contact.category:
                return False
        self.contacts.append(contact)
        self.save()
        return True

    def update(self, contact_id: str, save=True, **kwargs) -> bool:
        """연락처 수정"""
        for c in self.contacts:
            if c.id == contact_id:
                for key, value in kwargs.items():
                    if hasattr(c, key):
                        setattr(c, key, value)
                if save:
                    self.save()
                return True
        return False

    def batch_update_category(self, contact_ids: list, category: str):
        """여러 연락처 카테고리 일괄 변경 (한 번만 저장)"""
        id_set = set(contact_ids)
        for c in self.contacts:
            if c.id in id_set:
                c.category = category
        self.save()

    def delete(self, contact_id: str) -> bool:
        """연락처 삭제"""
        before = len(self.contacts)
        self.contacts = [c for c in self.contacts if c.id != contact_id]
        if len(self.contacts) < before:
            self.save()
            return True
        return False

    def get_by_category(self, category: str) -> list[Contact]:
        """카테고리별 연락처 조회"""
        if category == "all":
            return self.contacts
        return [c for c in self.contacts if c.category == category]

    def get_by_name(self, name: str) -> Optional[Contact]:
        """이름으로 연락처 검색"""
        for c in self.contacts:
            if c.name == name:
                return c
        return None

    def search(self, query: str) -> list[Contact]:
        """연락처 검색 (이름, 회사, 메모)"""
        query = query.lower()
        return [
            c for c in self.contacts
            if query in c.name.lower()
            or query in c.company.lower()
            or query in c.memo.lower()
        ]

    def get_all(self) -> list[Contact]:
        """전체 연락처 조회"""
        return self.contacts

    def get_count(self) -> int:
        """전체 연락처 수"""
        return len(self.contacts)

    def get_category_counts(self) -> dict:
        """카테고리별 연락처 수"""
        counts = {}
        for c in self.contacts:
            counts[c.category] = counts.get(c.category, 0) + 1
        return counts

    def import_from_excel(self, filepath: str, default_category: str = None) -> dict:
        """
        엑셀 파일에서 연락처 일괄 가져오기

        엑셀 형식:
        | 이름 | 카테고리 | 전화번호 | 회사 | 직급 | 메모 |

        Returns:
            {"success": int, "skipped": int, "errors": []}
        """
        if openpyxl is None:
            raise ImportError("openpyxl이 설치되지 않았습니다.")

        result = {"success": 0, "skipped": 0, "errors": []}

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active

            headers = []
            header_map = {
                "이름": "name", "name": "name",
                "카테고리": "category", "category": "category",
                "전화번호": "phone", "phone": "phone", "연락처": "phone",
                "회사": "company", "company": "company",
                "직급": "position", "position": "position",
                "메모": "memo", "memo": "memo", "비고": "memo",
                "생일": "birthday", "birthday": "birthday",
                "기념일": "anniversary", "anniversary": "anniversary",
            }

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # 헤더 행 파싱
                    headers = [
                        header_map.get(str(cell).strip().lower() if cell else "", "")
                        for cell in row
                    ]
                    continue

                if not row[0]:  # 이름이 없으면 건너뛰기
                    continue

                try:
                    data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers) and headers[col_idx]:
                            data[headers[col_idx]] = str(cell).strip() if cell else ""

                    if "name" not in data or not data["name"]:
                        result["skipped"] += 1
                        continue

                    # 카테고리: 엑셀에 있으면 사용, 없으면 default_category, 그것도 없으면 "미지정"
                    cat = data.get("category", "").strip()
                    if not cat:
                        cat = default_category or "미지정"
                    contact = Contact(
                        name=data.get("name", ""),
                        category=cat,
                        phone=data.get("phone", ""),
                        company=data.get("company", ""),
                        position=data.get("position", ""),
                        memo=data.get("memo", ""),
                        birthday=data.get("birthday", ""),
                        anniversary=data.get("anniversary", ""),
                    )

                    if self.add(contact):
                        result["success"] += 1
                    else:
                        result["skipped"] += 1

                except Exception as e:
                    result["errors"].append(f"행 {row_idx + 1}: {str(e)}")

            wb.close()

        except Exception as e:
            result["errors"].append(f"파일 읽기 오류: {str(e)}")

        return result

    def export_to_excel(self, filepath: str, contacts: list[Contact] = None) -> bool:
        """연락처를 엑셀 파일로 내보내기"""
        if openpyxl is None:
            raise ImportError("openpyxl이 설치되지 않았습니다.")

        contacts = contacts or self.contacts

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "연락처"

        # 헤더
        headers = ["이름", "카테고리", "전화번호", "회사", "직급", "메모",
                   "생일", "기념일", "마지막 발송", "발송 횟수"]
        ws.append(headers)

        # 데이터
        for c in contacts:
            ws.append([
                c.name, c.category, c.phone, c.company,
                c.position, c.memo, c.birthday or "", c.anniversary or "",
                c.last_sent or "", c.send_count
            ])

        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

        wb.save(filepath)
        return True

    def get_all_categories(self) -> list[str]:
        """기본 + 커스텀 + 실제 사용 중인 모든 카테고리 반환"""
        used = set(c.category for c in self.contacts)
        all_cats = list(self.DEFAULT_CATEGORIES)
        for cat in self.custom_categories:
            if cat not in all_cats:
                all_cats.append(cat)
        for cat in used:
            if cat not in all_cats:
                all_cats.append(cat)
        return all_cats

    def add_category(self, category: str) -> bool:
        """커스텀 카테고리 추가"""
        category = category.strip()
        if not category:
            return False
        if category in self.DEFAULT_CATEGORIES or category in self.custom_categories:
            return False
        self.custom_categories.append(category)
        self.save()
        return True

    def delete_category(self, category: str) -> bool:
        """커스텀 카테고리 삭제"""
        if category in self.custom_categories:
            self.custom_categories.remove(category)
            self.save()
            return True
        return False

    def create_sample_excel(self, filepath: str) -> bool:
        """샘플 엑셀 파일 생성 (가져오기 형식 안내)"""
        if openpyxl is None:
            raise ImportError("openpyxl이 설치되지 않았습니다.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "연락처"

        # 헤더
        headers = ["이름", "카테고리", "전화번호", "회사", "직급", "메모", "생일", "기념일"]
        ws.append(headers)

        # 헤더 스타일
        from openpyxl.styles import Font, PatternFill, Alignment
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 샘플 데이터
        samples = [
            ["홍길동", "friend", "010-1234-5678", "ABC회사", "대리", "동창", "03-15", ""],
            ["김영희", "business", "010-9876-5432", "XYZ건설", "팀장", "거래처 담당자", "07-22", "05-10"],
            ["이철수", "vip", "010-5555-1234", "DEF기획", "대표", "주요 고객", "", "12-25"],
            ["박지영", "family", "010-3333-4444", "", "", "사촌", "11-03", ""],
        ]
        for sample in samples:
            ws.append(sample)

        # 열 너비
        widths = [12, 12, 16, 16, 10, 20, 10, 10]
        for i, w in enumerate(widths, 1):
            col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
            ws.column_dimensions[col_letter].width = w

        # 안내 시트
        ws2 = wb.create_sheet("안내")
        ws2.append(["카카오톡 자동 발송 - 연락처 가져오기 안내"])
        ws2.append([])
        ws2.append(["필수 항목: 이름 (첫 번째 열)"])
        ws2.append(["선택 항목: 카테고리, 전화번호, 회사, 직급, 메모, 생일, 기념일"])
        ws2.append([])
        ws2.append(["생일/기념일 형식: MM-DD (예: 03-15, 12-25)"])
        ws2.append([])
        ws2.append(["카테고리 종류:"])
        all_cats = self.get_all_categories()
        ws2.append([", ".join(all_cats)])
        ws2.append([])
        ws2.append(["* 첫 번째 행은 반드시 헤더여야 합니다"])
        ws2.append(["* 이름이 비어있는 행은 건너뜁니다"])
        ws2.append(["* 동일 이름+카테고리 중복은 자동 건너뜁니다"])

        wb.save(filepath)
        return True

    def mark_sent(self, contact_id: str):
        """발송 완료 마킹"""
        for c in self.contacts:
            if c.id == contact_id:
                c.last_sent = datetime.now().isoformat()
                c.send_count += 1
                self.save()
                break
